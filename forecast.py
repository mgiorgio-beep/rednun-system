"""
Red Nun Revenue Forecast & Staffing Optimizer
Uses 12 weeks of Toast data to forecast next week's daily revenue
and compare against 7shifts scheduled labor.
"""
import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict

# Add parent dir for imports
sys.path.insert(0, os.path.dirname(__file__))

from data_store import get_connection

LOCATION_NAMES = {"dennis": "Dennis Port", "chatham": "Chatham"}

# Days Chatham / Dennis are normally closed
CLOSED_DAYS = {
    "chatham": [0],   # Monday
    "dennis": [0, 1], # Monday, Tuesday
}

# Target labor % by day type
TARGET_LABOR_PCT = {
    "slow": 30,   # Tue/Wed/Thu
    "busy": 25,   # Fri/Sat
    "moderate": 28, # Sun
}

DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def get_daily_history(location, weeks_back=12):
    """Get daily revenue history grouped by day of week."""
    conn = get_connection()
    cutoff = (datetime.now() - timedelta(weeks=weeks_back)).strftime("%Y%m%d")

    rows = conn.execute("""
        SELECT business_date,
               SUM(total_amount - tax_amount - tip_amount) as net_rev,
               COUNT(*) as orders
        FROM orders
        WHERE location = ? AND business_date >= ?
        GROUP BY business_date
        ORDER BY business_date
    """, (location, cutoff)).fetchall()
    conn.close()

    history = []
    for r in rows:
        bd = r["business_date"]
        dt = datetime.strptime(bd, "%Y%m%d")
        history.append({
            "date": bd,
            "dt": dt,
            "dow": dt.weekday(),
            "day_name": DAY_NAMES[dt.weekday()],
            "net_rev": r["net_rev"],
            "orders": r["orders"],
        })
    return history


def compute_dow_stats(history, location):
    """Compute day-of-week statistics from history."""
    closed = CLOSED_DAYS.get(location, [])
    by_dow = defaultdict(list)

    for d in history:
        if d["dow"] in closed:
            continue
        if d["net_rev"] < 200:  # Skip near-zero days
            continue
        by_dow[d["dow"]].append(d["net_rev"])

    stats = {}
    for dow in range(7):
        if dow in closed:
            stats[dow] = {"avg": 0, "median": 0, "low": 0, "high": 0,
                          "count": 0, "trend": 0, "closed": True}
            continue

        vals = by_dow.get(dow, [])
        if not vals:
            stats[dow] = {"avg": 0, "median": 0, "low": 0, "high": 0,
                          "count": 0, "trend": 0, "closed": False}
            continue

        vals_sorted = sorted(vals)
        n = len(vals_sorted)
        median = vals_sorted[n // 2]

        # Trend: compare last 4 weeks avg to overall avg
        recent = vals[-4:] if len(vals) >= 4 else vals
        recent_avg = sum(recent) / len(recent)
        overall_avg = sum(vals) / len(vals)
        trend_pct = ((recent_avg - overall_avg) / overall_avg * 100) if overall_avg > 0 else 0

        stats[dow] = {
            "avg": round(sum(vals) / len(vals)),
            "median": round(median),
            "low": round(min(vals)),
            "high": round(max(vals)),
            "count": len(vals),
            "trend": round(trend_pct, 1),
            "recent_avg": round(recent_avg),
            "closed": False,
        }
    return stats


def forecast_week(location, target_date=None):
    """
    Forecast next week's daily revenue.
    Uses weighted average: 60% recent 4-week trend, 40% overall DOW average.
    """
    if target_date is None:
        today = datetime.now().date()
        # Next Monday
        days_ahead = (7 - today.weekday()) % 7
        if days_ahead == 0:
            days_ahead = 7
        target_date = today + timedelta(days=days_ahead)

    history = get_daily_history(location)
    stats = compute_dow_stats(history, location)

    # Recent weekly revenue (last 4 operating weeks)
    recent_weeks = []
    for d in history:
        if d["net_rev"] < 200:
            continue
        wk = d["dt"].isocalendar()[1]
        yr = d["dt"].year
        key = f"{yr}-{wk}"
        if not recent_weeks or recent_weeks[-1]["key"] != key:
            recent_weeks.append({"key": key, "total": 0, "days": 0})
        recent_weeks[-1]["total"] += d["net_rev"]
        recent_weeks[-1]["days"] += 1

    # Trend multiplier from recent weeks
    if len(recent_weeks) >= 4:
        last4_avg = sum(w["total"] for w in recent_weeks[-4:]) / 4
        all_avg = sum(w["total"] for w in recent_weeks) / len(recent_weeks)
        week_trend = last4_avg / all_avg if all_avg > 0 else 1.0
    else:
        week_trend = 1.0

    forecast = []
    week_total = 0

    for i in range(7):
        day_date = target_date + timedelta(days=i)
        dow = day_date.weekday()
        s = stats.get(dow, {})

        if s.get("closed", False) or s.get("avg", 0) == 0:
            forecast.append({
                "date": day_date.strftime("%Y%m%d"),
                "day_name": DAY_NAMES[dow],
                "forecast": 0,
                "low": 0,
                "high": 0,
                "confidence": "closed",
            })
            continue

        # Weighted forecast: 60% recent trend, 40% historical avg
        recent_avg = s.get("recent_avg", s["avg"])
        base_forecast = (recent_avg * 0.6 + s["avg"] * 0.4)

        # Apply overall week trend
        adjusted = base_forecast * min(max(week_trend, 0.7), 1.3)

        # Confidence based on sample size and variance
        if s["count"] >= 8:
            spread = (s["high"] - s["low"]) / s["avg"] if s["avg"] > 0 else 1
            confidence = "high" if spread < 0.6 else "medium"
        else:
            confidence = "low"

        low_est = round(adjusted * 0.85)
        high_est = round(adjusted * 1.15)

        forecast.append({
            "date": day_date.strftime("%Y%m%d"),
            "day_name": DAY_NAMES[dow],
            "forecast": round(adjusted),
            "low": low_est,
            "high": high_est,
            "confidence": confidence,
            "hist_avg": s["avg"],
            "trend": s["trend"],
        })
        week_total += adjusted

    return {
        "location": location,
        "week_start": target_date.strftime("%Y%m%d"),
        "week_total": round(week_total),
        "daily": forecast,
        "week_trend": round((week_trend - 1) * 100, 1),
    }


def get_labor_forecast_comparison(location, target_date=None):
    """
    Compare forecasted revenue against scheduled labor.
    Returns staffing recommendations per day.
    """
    fc = forecast_week(location, target_date)

    try:
        from sevenshifts_client import get_labor_for_report
        start = fc["week_start"]
        end_dt = datetime.strptime(start, "%Y%m%d") + timedelta(days=6)
        end = end_dt.strftime("%Y%m%d")
        labor = get_labor_for_report(location, start, end)
    except Exception:
        labor = None

    recommendations = []
    for day in fc["daily"]:
        rec = {**day, "labor_cost": 0, "labor_pct": 0, "status": "", "action": ""}

        if day["confidence"] == "closed":
            rec["status"] = "CLOSED"
            recommendations.append(rec)
            continue

        rev = day["forecast"]

        # Determine target labor %
        dow = datetime.strptime(day["date"], "%Y%m%d").weekday()
        if dow in (4, 5):  # Fri, Sat
            target = TARGET_LABOR_PCT["busy"]
        elif dow == 6:  # Sun
            target = TARGET_LABOR_PCT["moderate"]
        else:
            target = TARGET_LABOR_PCT["slow"]

        target_labor = rev * target / 100
        rec["target_labor"] = round(target_labor)
        rec["target_pct"] = target

        if rev < 1500:
            rec["status"] = "⚠️ SLOW"
            rec["action"] = f"Skeleton crew. Target labor ${target_labor:,.0f} ({target}%)"
        elif rev < 3000:
            rec["status"] = "📊 MODERATE"
            rec["action"] = f"Standard crew. Target labor ${target_labor:,.0f} ({target}%)"
        else:
            rec["status"] = "🔥 BUSY"
            rec["action"] = f"Full staff. Target labor ${target_labor:,.0f} ({target}%)"

        recommendations.append(rec)

    return {
        "location": location,
        "forecast": fc,
        "recommendations": recommendations,
    }


def print_forecast_report(target_date=None):
    """Print formatted forecast report for both locations."""

    print("=" * 70)
    print("  RED NUN — WEEKLY REVENUE FORECAST & STAFFING GUIDE")
    print("=" * 70)

    for loc in ["dennis", "chatham"]:
        result = get_labor_forecast_comparison(loc, target_date)
        fc = result["forecast"]
        recs = result["recommendations"]
        name = LOCATION_NAMES[loc]

        print(f"\n{'─' * 70}")
        print(f"  {name}")
        print(f"  Week of {fc['week_start']}  |  Forecast: ${fc['week_total']:,}")
        trend_dir = "↑" if fc["week_trend"] > 0 else "↓" if fc["week_trend"] < 0 else "→"
        print(f"  4-Week Trend: {trend_dir} {abs(fc['week_trend'])}%")
        print(f"{'─' * 70}")
        print(f"  {'Day':<5} {'Date':<10} {'Forecast':>10} {'Range':>18} {'Status':<12} {'Target Labor':>14}")
        print(f"  {'─'*5} {'─'*10} {'─'*10} {'─'*18} {'─'*12} {'─'*14}")

        for r in recs:
            if r.get("confidence") == "closed":
                print(f"  {r['day_name']:<5} {r['date']:<10} {'—':>10} {'':>18} {'CLOSED':<12}")
                continue

            range_str = f"${r['low']:,} – ${r['high']:,}"
            target_str = f"${r.get('target_labor', 0):,} ({r.get('target_pct', 0)}%)"
            print(f"  {r['day_name']:<5} {r['date']:<10} ${r['forecast']:>8,} {range_str:>18} {r['status']:<12} {target_str:>14}")

    # Day-of-week patterns
    print(f"\n{'=' * 70}")
    print("  HISTORICAL DAY-OF-WEEK PATTERNS (12 weeks)")
    print(f"{'=' * 70}")

    for loc in ["dennis", "chatham"]:
        name = LOCATION_NAMES[loc]
        history = get_daily_history(loc)
        stats = compute_dow_stats(history, loc)

        print(f"\n  {name}:")
        print(f"  {'Day':<5} {'Avg':>8} {'Median':>8} {'Low':>8} {'High':>8} {'#Wks':>6} {'Trend':>8}")
        print(f"  {'─'*5} {'─'*8} {'─'*8} {'─'*8} {'─'*8} {'─'*6} {'─'*8}")

        for dow in range(7):
            s = stats[dow]
            if s.get("closed"):
                print(f"  {DAY_NAMES[dow]:<5} {'CLOSED':>8}")
                continue
            if s["count"] == 0:
                continue

            trend_str = f"{'+' if s['trend'] > 0 else ''}{s['trend']}%"
            print(f"  {DAY_NAMES[dow]:<5} ${s['avg']:>6,} ${s['median']:>6,} ${s['low']:>6,} ${s['high']:>6,} {s['count']:>6} {trend_str:>8}")


def generate_forecast_excel(target_date=None, output_dir=None):
    """Generate Excel forecast workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(__file__), "exports")
    os.makedirs(output_dir, exist_ok=True)

    if target_date is None:
        today = datetime.now().date()
        days_ahead = (7 - today.weekday()) % 7
        if days_ahead == 0:
            days_ahead = 7
        target_date = today + timedelta(days=days_ahead)

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Forecast"

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2D3142")
    money_fmt = "$#,##0"
    pct_fmt = "0.0%"
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD")
    )
    bold_font = Font(name="Arial", bold=True, size=11)
    red_font = Font(name="Arial", bold=True, color="CC0000")
    green_font = Font(name="Arial", bold=True, color="228B22")

    slow_fill = PatternFill("solid", fgColor="FFF3CD")
    mod_fill = PatternFill("solid", fgColor="D1ECF1")
    busy_fill = PatternFill("solid", fgColor="D4EDDA")

    row = 1
    for loc in ["dennis", "chatham"]:
        result = get_labor_forecast_comparison(loc, target_date)
        fc = result["forecast"]
        recs = result["recommendations"]
        name = LOCATION_NAMES[loc]

        # Location header
        loc_fill = PatternFill("solid", fgColor="C43B3B" if loc == "dennis" else "4A7FD4")
        ws.cell(row, 1, f"{name} — Week of {fc['week_start']}").font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
        ws.cell(row, 1).fill = loc_fill
        for c in range(1, 8):
            ws.cell(row, c).fill = loc_fill
        row += 1

        ws.cell(row, 1, f"Forecast Total: ${fc['week_total']:,}").font = bold_font
        trend_dir = "↑" if fc["week_trend"] > 0 else "↓"
        ws.cell(row, 3, f"4-Week Trend: {trend_dir} {abs(fc['week_trend'])}%").font = bold_font
        row += 1

        # Column headers
        headers = ["Day", "Date", "Forecast", "Low", "High", "Status", "Target Labor"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for r in recs:
            ws.cell(row, 1, r["day_name"]).font = Font(name="Arial", size=10)
            ws.cell(row, 2, r["date"]).font = Font(name="Arial", size=10)

            if r.get("confidence") == "closed":
                ws.cell(row, 3, "CLOSED").font = Font(name="Arial", italic=True, color="999999")
                row += 1
                continue

            ws.cell(row, 3, r["forecast"]).number_format = money_fmt
            ws.cell(row, 4, r["low"]).number_format = money_fmt
            ws.cell(row, 5, r["high"]).number_format = money_fmt
            ws.cell(row, 6, r.get("status", ""))

            target = r.get("target_labor", 0)
            ws.cell(row, 7, target).number_format = money_fmt

            # Color code by status
            if "SLOW" in r.get("status", ""):
                for c in range(1, 8):
                    ws.cell(row, c).fill = slow_fill
            elif "BUSY" in r.get("status", ""):
                for c in range(1, 8):
                    ws.cell(row, c).fill = busy_fill
            else:
                for c in range(1, 8):
                    ws.cell(row, c).fill = mod_fill

            for c in range(1, 8):
                ws.cell(row, c).border = thin_border
            row += 1

        row += 1  # Spacer

    # Historical patterns sheet
    ws2 = wb.create_sheet("DOW Patterns")
    row = 1
    for loc in ["dennis", "chatham"]:
        name = LOCATION_NAMES[loc]
        history = get_daily_history(loc)
        stats = compute_dow_stats(history, loc)

        loc_fill = PatternFill("solid", fgColor="C43B3B" if loc == "dennis" else "4A7FD4")
        ws2.cell(row, 1, f"{name} — 12 Week Patterns").font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
        for c in range(1, 8):
            ws2.cell(row, c).fill = loc_fill
        row += 1

        headers = ["Day", "Average", "Median", "Low", "High", "Weeks", "Trend"]
        for c, h in enumerate(headers, 1):
            cell = ws2.cell(row, c, h)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        for dow in range(7):
            s = stats[dow]
            if s.get("closed"):
                ws2.cell(row, 1, DAY_NAMES[dow])
                ws2.cell(row, 2, "CLOSED").font = Font(name="Arial", italic=True, color="999999")
                row += 1
                continue
            if s["count"] == 0:
                continue

            ws2.cell(row, 1, DAY_NAMES[dow])
            ws2.cell(row, 2, s["avg"]).number_format = money_fmt
            ws2.cell(row, 3, s["median"]).number_format = money_fmt
            ws2.cell(row, 4, s["low"]).number_format = money_fmt
            ws2.cell(row, 5, s["high"]).number_format = money_fmt
            ws2.cell(row, 6, s["count"])
            ws2.cell(row, 7, s["trend"] / 100).number_format = "0.0%"

            if s["trend"] > 5:
                ws2.cell(row, 7).font = green_font
            elif s["trend"] < -5:
                ws2.cell(row, 7).font = red_font

            for c in range(1, 8):
                ws2.cell(row, c).border = thin_border
            row += 1
        row += 1

    # Auto-width
    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_sheet.column_dimensions[col_letter].width = min(max_len + 3, 20)

    wk = target_date.strftime("%Y%m%d")
    filepath = os.path.join(output_dir, f"RedNun_Forecast_{wk}.xlsx")
    wb.save(filepath)
    print(f"Saved: {filepath}")
    return filepath


if __name__ == "__main__":
    target = None
    if len(sys.argv) == 2:
        target = datetime.strptime(sys.argv[1], "%Y%m%d").date()

    print_forecast_report(target)
    print()
    generate_forecast_excel(target)
