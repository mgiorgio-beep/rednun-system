"""
Vendor Payment Tracking routes — import, list, export vendor payments.
Blueprint: payment_bp at /api/payments/*

Tracks payment batches (ACH transfers, checks, etc.) from vendor portals,
each covering one or more invoices. Cross-links to scanned_invoices for
payment status tracking.
"""

import io
import json
import logging
import os
import subprocess
import threading
from collections import defaultdict
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file

from data_store import get_connection

logger = logging.getLogger(__name__)

payment_bp = Blueprint("payment_bp", __name__)

# ─── TABLE INIT ──────────────────────────────────────────────────────────────


def init_payment_tables():
    """Create vendor_payments and vendor_payment_invoices tables if they don't exist."""
    conn = get_connection()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS vendor_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor TEXT NOT NULL,
            location TEXT,
            payment_date TEXT NOT NULL,
            payment_ref TEXT UNIQUE,
            payment_method TEXT,
            payment_total REAL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_vp_vendor ON vendor_payments(vendor);
        CREATE INDEX IF NOT EXISTS idx_vp_date ON vendor_payments(payment_date);
        CREATE INDEX IF NOT EXISTS idx_vp_ref ON vendor_payments(payment_ref);

        CREATE TABLE IF NOT EXISTS vendor_payment_invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            payment_id INTEGER NOT NULL,
            invoice_number TEXT,
            invoice_date TEXT,
            due_date TEXT,
            amount_paid REAL DEFAULT 0,
            FOREIGN KEY (payment_id) REFERENCES vendor_payments(id)
        );

        CREATE INDEX IF NOT EXISTS idx_vpi_payment ON vendor_payment_invoices(payment_id);
        CREATE INDEX IF NOT EXISTS idx_vpi_invoice ON vendor_payment_invoices(invoice_number);
    """)

    # ── New columns for centralized payments ──
    migrations = [
        "ALTER TABLE vendor_payments ADD COLUMN check_number TEXT",
        "ALTER TABLE vendor_payments ADD COLUMN memo TEXT",
        "ALTER TABLE vendor_payments ADD COLUMN status TEXT DEFAULT 'cleared'",
        "ALTER TABLE vendor_payments ADD COLUMN source TEXT DEFAULT 'import'",
        "ALTER TABLE vendor_payments ADD COLUMN updated_at TEXT",
        "ALTER TABLE vendor_payments ADD COLUMN ap_payment_id INTEGER",
        # Portal pay setting on vendor_bill_pay
        "ALTER TABLE vendor_bill_pay ADD COLUMN portal_pay_enabled INTEGER DEFAULT 0",
    ]
    for sql in migrations:
        try:
            conn.execute(sql)
        except Exception:
            pass  # column already exists

    # ── Backfill existing ap_payments into vendor_payments ──
    try:
        ap_rows = conn.execute("SELECT * FROM ap_payments").fetchall()
        for ap in ap_rows:
            existing = conn.execute(
                "SELECT id FROM vendor_payments WHERE ap_payment_id = ?",
                (ap["id"],),
            ).fetchone()
            if existing:
                continue
            ref = f"CHK-{ap['check_number']}" if ap.get("check_number") else f"CHK-AP{ap['id']}"
            status_map = {"void": "void", "printed": "printed", "pending": "pending", "cleared": "cleared"}
            vp_status = status_map.get(ap["status"], "pending")
            try:
                cur = conn.execute(
                    """INSERT INTO vendor_payments
                       (vendor, location, payment_date, payment_ref, payment_method,
                        payment_total, check_number, memo, status, source, ap_payment_id)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'check', ?)""",
                    (
                        ap["vendor_name"], None, ap["payment_date"], ref,
                        ap.get("payment_method", "check"), ap["amount"],
                        ap.get("check_number"), ap.get("memo"), vp_status, ap["id"],
                    ),
                )
                vp_id = cur.lastrowid
                # Mirror invoice links
                links = conn.execute(
                    """SELECT pi.amount_applied, si.invoice_number, si.invoice_date, si.due_date
                       FROM ap_payment_invoices pi
                       JOIN scanned_invoices si ON si.id = pi.invoice_id
                       WHERE pi.payment_id = ?""",
                    (ap["id"],),
                ).fetchall()
                for link in links:
                    conn.execute(
                        """INSERT INTO vendor_payment_invoices
                           (payment_id, invoice_number, invoice_date, due_date, amount_paid)
                           VALUES (?, ?, ?, ?, ?)""",
                        (vp_id, link["invoice_number"], link["invoice_date"],
                         link["due_date"], link["amount_applied"]),
                    )
            except Exception as e:
                logger.debug(f"Backfill skip ap_payment {ap['id']}: {e}")
        conn.commit()
    except Exception:
        pass  # ap_payments table may not exist yet

    conn.close()
    logger.info("Vendor payment tables initialized")


# ─── HELPERS ─────────────────────────────────────────────────────────────────

LOCATION_MAP = {
    "red nun dennisport": "Dennis",
    "dennisport": "Dennis",
    "dennis port": "Dennis",
    "dennis": "Dennis",
    "red nun chatham": "Chatham",
    "chatham": "Chatham",
}


def detect_location(customer_name):
    """Detect location from customer name string."""
    lower = (customer_name or "").lower()
    for key, loc in LOCATION_MAP.items():
        if key in lower:
            return loc
    return None


def _get_summary(conn, where_clause="", params=()):
    """Build summary stats, optionally filtered."""
    base = f"FROM vendor_payments {where_clause}" if where_clause else "FROM vendor_payments"

    total_all = conn.execute(
        f"SELECT COALESCE(SUM(payment_total), 0) {base}", params
    ).fetchone()[0]

    # This month
    month_start = datetime.now().strftime("%Y-%m-01")
    if where_clause:
        month_sql = f"SELECT COALESCE(SUM(payment_total), 0) FROM vendor_payments {where_clause} AND payment_date >= ?"
        month_params = params + (month_start,)
    else:
        month_sql = "SELECT COALESCE(SUM(payment_total), 0) FROM vendor_payments WHERE payment_date >= ?"
        month_params = (month_start,)
    total_month = conn.execute(month_sql, month_params).fetchone()[0]

    # By vendor
    by_vendor_rows = conn.execute(
        f"SELECT vendor, location, SUM(payment_total) as total, COUNT(*) as count "
        f"{base} GROUP BY vendor, location ORDER BY total DESC",
        params,
    ).fetchall()
    by_vendor = [
        {"vendor": r["vendor"], "location": r["location"] or "", "total": r["total"], "count": r["count"]}
        for r in by_vendor_rows
    ]

    return {
        "total_all_time": round(total_all, 2),
        "total_this_month": round(total_month, 2),
        "by_vendor": by_vendor,
    }


def _cross_link_invoice(conn, invoice_number, payment_ref, paid_date):
    """Try to update scanned_invoices payment_status when a payment references an invoice."""
    if not invoice_number:
        return
    conn.execute(
        """UPDATE scanned_invoices
           SET payment_status = 'paid',
               payment_reference = ?,
               paid_date = ?
           WHERE invoice_number = ? COLLATE NOCASE
             AND status = 'confirmed'
             AND (payment_status IS NULL OR payment_status = 'unpaid')""",
        (payment_ref, paid_date, str(invoice_number)),
    )


def _import_payments(payments_list):
    """Core import logic shared by /import and /upload. Returns (imported, skipped)."""
    conn = get_connection()
    imported = 0
    skipped = 0

    for p in payments_list:
        vendor = p.get("vendor", "")
        location = p.get("location", "")
        payment_date = p.get("payment_date", "")
        payment_ref = p.get("payment_ref", "")
        payment_method = p.get("payment_method", "")
        payment_total = float(p.get("payment_total", 0))
        invoices = p.get("invoices", [])

        if not payment_ref:
            skipped += 1
            continue

        try:
            cur = conn.execute(
                """INSERT OR IGNORE INTO vendor_payments
                   (vendor, location, payment_date, payment_ref, payment_method, payment_total)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (vendor, location, payment_date, payment_ref, payment_method, payment_total),
            )
            if cur.rowcount == 0:
                skipped += 1
                continue

            payment_id = cur.lastrowid
            for inv in invoices:
                conn.execute(
                    """INSERT INTO vendor_payment_invoices
                       (payment_id, invoice_number, invoice_date, due_date, amount_paid)
                       VALUES (?, ?, ?, ?, ?)""",
                    (
                        payment_id,
                        str(inv.get("invoice_number", "")),
                        inv.get("invoice_date", ""),
                        inv.get("due_date", ""),
                        float(inv.get("amount_paid", 0)),
                    ),
                )
                _cross_link_invoice(
                    conn,
                    inv.get("invoice_number"),
                    payment_ref,
                    payment_date,
                )

            imported += 1
        except Exception as e:
            logger.warning(f"Payment import error for ref={payment_ref}: {e}")
            skipped += 1

    conn.commit()
    conn.close()
    return imported, skipped


# ─── XLSX PARSER ─────────────────────────────────────────────────────────────


def parse_usfoods_payment_xlsx(file_stream):
    """Parse US Foods payment xlsx into grouped payment dicts.

    Expected columns (case-insensitive):
        Customer Name, Document Date, Primary Transaction Number,
        Paid Amount, Payment Date, Reference Number, Payment Method
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active

    # Map expected headers to column indices
    header_map = {
        "customer name": "customer_name",
        "document date": "document_date",
        "primary transaction number": "invoice_number",
        "paid amount": "amount_paid",
        "payment date": "payment_date",
        "reference number": "payment_ref",
        "payment method": "payment_method",
    }

    col_idx = {}
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=False), 1):
        for cell in row:
            val = str(cell.value or "").strip().lower()
            if val in header_map:
                col_idx[header_map[val]] = cell.column - 1
        if len(col_idx) >= 4:
            header_row = row_idx
            break

    if not header_row or "payment_ref" not in col_idx:
        wb.close()
        return []

    # Read data rows and group by reference number
    groups = defaultdict(list)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row)

        def get(key, default=""):
            idx = col_idx.get(key)
            if idx is not None and idx < len(vals) and vals[idx] is not None:
                return vals[idx]
            return default

        ref = str(get("payment_ref", "")).strip()
        if not ref:
            continue
        groups[ref].append({
            "customer_name": str(get("customer_name", "")),
            "document_date": _fmt_date(get("document_date")),
            "invoice_number": str(get("invoice_number", "")),
            "amount_paid": _to_float(get("amount_paid", 0)),
            "payment_date": _fmt_date(get("payment_date")),
            "payment_method": str(get("payment_method", "")),
        })

    wb.close()

    # Build payment objects from groups
    payments = []
    for ref, rows in groups.items():
        first = rows[0]
        customer = first["customer_name"]
        location = detect_location(customer)

        # Derive vendor from customer name
        vendor = "US Foods"
        if "pfg" in customer.lower() or "performance" in customer.lower():
            vendor = "PFG"

        payment_total = sum(r["amount_paid"] for r in rows)
        invoices = [
            {
                "invoice_number": r["invoice_number"],
                "invoice_date": r["document_date"],
                "amount_paid": r["amount_paid"],
            }
            for r in rows
        ]

        payments.append({
            "vendor": vendor,
            "location": location or "",
            "payment_date": first["payment_date"],
            "payment_ref": ref,
            "payment_method": first["payment_method"],
            "payment_total": round(payment_total, 2),
            "invoices": invoices,
        })

    return payments


def _fmt_date(val):
    """Normalize a date value to YYYY-MM-DD string."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val or "").strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _to_float(val):
    """Convert a value to float, stripping currency symbols."""
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val or "0").replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


# ─── ENDPOINTS ───────────────────────────────────────────────────────────────


@payment_bp.route("/api/payments", methods=["GET"])
def api_list_payments():
    """List all payments with nested invoice details and summary stats."""
    vendor = request.args.get("vendor")
    location = request.args.get("location")
    source = request.args.get("source")
    status = request.args.get("status")

    conn = get_connection()

    # Build WHERE clause
    conditions = []
    params = []
    if vendor:
        conditions.append("vendor = ?")
        params.append(vendor)
    if location:
        conditions.append("location = ?")
        params.append(location)
    if source:
        conditions.append("source = ?")
        params.append(source)
    if status:
        conditions.append("status = ?")
        params.append(status)

    where = ""
    if conditions:
        where = "WHERE " + " AND ".join(conditions)

    # Fetch payments
    rows = conn.execute(
        f"SELECT * FROM vendor_payments {where} ORDER BY payment_date DESC, id DESC",
        params,
    ).fetchall()

    payments = []
    for r in rows:
        inv_rows = conn.execute(
            "SELECT * FROM vendor_payment_invoices WHERE payment_id = ? ORDER BY id",
            (r["id"],),
        ).fetchall()
        payments.append({
            "id": r["id"],
            "vendor": r["vendor"],
            "location": r["location"],
            "payment_date": r["payment_date"],
            "payment_ref": r["payment_ref"],
            "payment_method": r["payment_method"],
            "payment_total": r["payment_total"],
            "check_number": r["check_number"],
            "memo": r["memo"],
            "status": r["status"] or "cleared",
            "source": r["source"] or "import",
            "ap_payment_id": r["ap_payment_id"],
            "created_at": r["created_at"],
            "invoices": [
                {
                    "id": i["id"],
                    "invoice_number": i["invoice_number"],
                    "invoice_date": i["invoice_date"],
                    "due_date": i["due_date"],
                    "amount_paid": i["amount_paid"],
                }
                for i in inv_rows
            ],
        })

    summary = _get_summary(conn, where, tuple(params))
    conn.close()

    return jsonify({"payments": payments, "count": len(payments), "summary": summary})


@payment_bp.route("/api/payments/<int:payment_id>/void", methods=["PUT"])
def api_void_payment(payment_id):
    """Void a vendor_payment. If it has ap_payment_id, also void that."""
    conn = get_connection()
    vp = conn.execute("SELECT * FROM vendor_payments WHERE id = ?", (payment_id,)).fetchone()
    if not vp:
        conn.close()
        return jsonify({"error": "Payment not found"}), 404
    if vp["status"] == "void":
        conn.close()
        return jsonify({"error": "Already voided"}), 400

    now = datetime.now().isoformat()
    conn.execute(
        "UPDATE vendor_payments SET status = 'void', updated_at = ? WHERE id = ?",
        (now, payment_id),
    )

    # If linked to an ap_payment, void that too + reverse invoice balances
    ap_id = vp["ap_payment_id"]
    if ap_id:
        ap = conn.execute("SELECT * FROM ap_payments WHERE id = ?", (ap_id,)).fetchone()
        if ap and ap["status"] != "void":
            links = conn.execute(
                "SELECT invoice_id, amount_applied FROM ap_payment_invoices WHERE payment_id = ?",
                (ap_id,),
            ).fetchall()
            for link in links:
                conn.execute("""
                    UPDATE scanned_invoices
                    SET amount_paid = MAX(0, COALESCE(amount_paid, 0) - ?),
                        balance = COALESCE(balance, 0) + ?,
                        payment_status = CASE
                            WHEN COALESCE(amount_paid, 0) - ? <= 0 THEN 'unpaid'
                            ELSE 'partial'
                        END
                    WHERE id = ?
                """, (link["amount_applied"], link["amount_applied"],
                      link["amount_applied"], link["invoice_id"]))
            conn.execute(
                "UPDATE ap_payments SET status = 'void', updated_at = ? WHERE id = ?",
                (now, ap_id),
            )

    conn.commit()
    conn.close()
    logger.info(f"Vendor payment #{payment_id} voided")
    return jsonify({"status": "ok"})


@payment_bp.route("/api/payments/<int:payment_id>", methods=["DELETE"])
def api_delete_payment(payment_id):
    """Delete a vendor_payment record entirely. No invoice balance reversal.
    Use for failed/processing portal payments that never actually paid anything."""
    conn = get_connection()
    vp = conn.execute("SELECT * FROM vendor_payments WHERE id = ?", (payment_id,)).fetchone()
    if not vp:
        conn.close()
        return jsonify({"error": "Payment not found"}), 404

    # Delete invoice links first, then the payment
    conn.execute("DELETE FROM vendor_payment_invoices WHERE payment_id = ?", (payment_id,))
    conn.execute("DELETE FROM vendor_payments WHERE id = ?", (payment_id,))
    conn.commit()
    conn.close()
    logger.info(f"Vendor payment #{payment_id} deleted (source={vp['source']}, status={vp['status']})")
    return jsonify({"status": "ok"})


@payment_bp.route("/api/payments/import", methods=["POST"])
def api_import_payments():
    """Bulk import payments from JSON. Dedup by payment_ref."""
    data = request.get_json(silent=True) or {}
    payments_list = data.get("payments", [])

    if not payments_list:
        return jsonify({"error": "No payments provided"}), 400

    imported, skipped = _import_payments(payments_list)
    return jsonify({
        "status": "ok",
        "imported": imported,
        "skipped_duplicates": skipped,
    })


@payment_bp.route("/api/payments/upload", methods=["POST"])
def api_upload_payments():
    """Upload xlsx or PDF file with payment data."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]
    fname = (f.filename or "").lower()

    if fname.endswith(".xlsx"):
        try:
            payments = parse_usfoods_payment_xlsx(f.stream)
            if not payments:
                return jsonify({"error": "Could not parse xlsx — no payment data found. Check column headers."}), 400
            imported, skipped = _import_payments(payments)
            return jsonify({
                "status": "ok",
                "format": "xlsx",
                "imported": imported,
                "skipped_duplicates": skipped,
                "total_parsed": len(payments),
            })
        except Exception as e:
            logger.exception("xlsx upload parse error")
            return jsonify({"error": f"Failed to parse xlsx: {e}"}), 400

    elif fname.endswith(".pdf"):
        upload_dir = os.path.join(os.path.dirname(__file__), "data", "payment_uploads")
        os.makedirs(upload_dir, exist_ok=True)
        dest = os.path.join(upload_dir, f.filename or "payment.pdf")
        f.save(dest)
        return jsonify({
            "status": "ok",
            "format": "pdf",
            "message": "PDF saved for manual review",
            "path": dest,
        })

    else:
        return jsonify({"error": "Unsupported file type. Please upload .xlsx or .pdf"}), 400


@payment_bp.route("/api/payments/export", methods=["GET"])
def api_export_payments():
    """Export payments as an Excel file grouped by vendor."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    vendor = request.args.get("vendor")
    location = request.args.get("location")

    conn = get_connection()

    conditions = []
    params = []
    if vendor:
        conditions.append("vendor = ?")
        params.append(vendor)
    if location:
        conditions.append("location = ?")
        params.append(location)

    where = ""
    if conditions:
        where = "WHERE " + " AND ".join(conditions)

    rows = conn.execute(
        f"SELECT * FROM vendor_payments {where} ORDER BY vendor, location, payment_date DESC",
        params,
    ).fetchall()

    # Group by vendor+location
    groups = defaultdict(list)
    for r in rows:
        key = f"{r['vendor']} {r['location'] or ''}".strip()
        inv_rows = conn.execute(
            "SELECT * FROM vendor_payment_invoices WHERE payment_id = ? ORDER BY id",
            (r["id"],),
        ).fetchall()
        groups[key].append({"payment": r, "invoices": inv_rows})

    conn.close()

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="1e293b")
    header_font_white = Font(bold=True, color="e2e8f0", size=11)
    thin_border = Border(bottom=Side(style="thin", color="334155"))
    money_fmt = '#,##0.00'

    current_row = 1

    # Title
    ws.cell(row=current_row, column=1, value="Red Nun — Vendor Payments Report")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
    current_row += 1
    ws.cell(row=current_row, column=1, value=f"Generated {datetime.now().strftime('%m/%d/%Y %I:%M %p')}")
    ws.cell(row=current_row, column=1).font = Font(size=10, italic=True)
    current_row += 2

    grand_total = 0

    for group_name, entries in groups.items():
        # Vendor header
        ws.cell(row=current_row, column=1, value=group_name)
        ws.cell(row=current_row, column=1).font = header_font
        current_row += 1

        # Column headers
        headers = ["Payment Date", "Payment Ref", "Method", "Invoice #", "Invoice Date", "Amount"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=ci, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
        current_row += 1

        group_total = 0
        for entry in entries:
            p = entry["payment"]
            invs = entry["invoices"]
            group_total += p["payment_total"]

            # Payment summary row
            ws.cell(row=current_row, column=1, value=p["payment_date"])
            ws.cell(row=current_row, column=2, value=p["payment_ref"])
            ws.cell(row=current_row, column=3, value=p["payment_method"])
            ws.cell(row=current_row, column=6, value=p["payment_total"])
            ws.cell(row=current_row, column=6).number_format = money_fmt
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=6).font = Font(bold=True)
            current_row += 1

            # Invoice detail rows
            for inv in invs:
                ws.cell(row=current_row, column=4, value=inv["invoice_number"])
                ws.cell(row=current_row, column=5, value=inv["invoice_date"])
                ws.cell(row=current_row, column=6, value=inv["amount_paid"])
                ws.cell(row=current_row, column=6).number_format = money_fmt
                ws.cell(row=current_row, column=4).font = Font(size=10)
                ws.cell(row=current_row, column=5).font = Font(size=10)
                ws.cell(row=current_row, column=6).font = Font(size=10)
                current_row += 1

        # Group total
        ws.cell(row=current_row, column=5, value=f"{group_name} Total:")
        ws.cell(row=current_row, column=5).font = Font(bold=True)
        ws.cell(row=current_row, column=6, value=group_total)
        ws.cell(row=current_row, column=6).number_format = money_fmt
        ws.cell(row=current_row, column=6).font = Font(bold=True)
        current_row += 2

        grand_total += group_total

    # Grand total
    ws.cell(row=current_row, column=5, value="Grand Total:")
    ws.cell(row=current_row, column=5).font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=6, value=grand_total)
    ws.cell(row=current_row, column=6).number_format = money_fmt
    ws.cell(row=current_row, column=6).font = Font(bold=True, size=12)

    # Set column widths
    widths = [14, 24, 14, 16, 14, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Red_Nun_Payments_{datetime.now().strftime('%Y%m%d')}.xlsx",
    )


# ─── PAYMENT SCRAPER FRAMEWORK ───────────────────────────────────────────────

_PAYMENT_SCRAPER_REGISTRY = {
    "usfoods": {
        "display_name": "US Foods",
        "script_dir": os.path.expanduser("~/vendor-scrapers/usfoods-payments"),
        "script": "usfoods_payment_scraper.py",
    },
}

_VENDOR_KEY_MAP = {
    "US Foods": "usfoods",
}

_PYTHON = "/opt/rednun/venv/bin/python3"
_PAYMENT_LOG_DIR = os.path.expanduser("~/vendor-scrapers/logs")
_PAYMENT_STATE_FILE = os.path.expanduser("~/vendor-scrapers/logs/payment_scraper_state.json")
_payment_lock = threading.Lock()


def _read_payment_scraper_state():
    try:
        if os.path.exists(_PAYMENT_STATE_FILE):
            with open(_PAYMENT_STATE_FILE) as f:
                return json.load(f)
    except Exception:
        pass
    return {"running": {}, "results": {}}


def _write_payment_scraper_state(state):
    os.makedirs(os.path.dirname(_PAYMENT_STATE_FILE), exist_ok=True)
    tmp = _PAYMENT_STATE_FILE + ".tmp"
    with open(tmp, "w") as f:
        json.dump(state, f)
    os.replace(tmp, _PAYMENT_STATE_FILE)


def _set_payment_running(key, display_name):
    with _payment_lock:
        state = _read_payment_scraper_state()
        state["running"][key] = {"started": datetime.now().isoformat(), "display_name": display_name}
        _write_payment_scraper_state(state)


def _clear_payment_running(key):
    with _payment_lock:
        state = _read_payment_scraper_state()
        state["running"].pop(key, None)
        _write_payment_scraper_state(state)


def _set_payment_result(key, exit_code, tail):
    with _payment_lock:
        state = _read_payment_scraper_state()
        state["results"][key] = {
            "exit_code": exit_code,
            "tail": tail,
            "finished": datetime.now().isoformat(),
        }
        _write_payment_scraper_state(state)


def _run_payment_scraper_bg(key, vendor_payment_id, scraper_info):
    """Run payment scraper in background thread. Update vendor_payment on completion."""
    os.makedirs(_PAYMENT_LOG_DIR, exist_ok=True)
    log_path = os.path.join(_PAYMENT_LOG_DIR, f"payment_{key}.log")
    display_name = scraper_info["display_name"]
    _set_payment_running(key, display_name)

    try:
        result = subprocess.run(
            [_PYTHON, scraper_info["script"]],
            cwd=scraper_info["script_dir"],
            capture_output=True, text=True, timeout=300,
        )
        with open(log_path, "w") as f:
            f.write(result.stdout or "")
            if result.stderr:
                f.write("\n--- STDERR ---\n")
                f.write(result.stderr)

        all_output = (result.stdout or "") + (result.stderr or "")
        tail = "\n".join(all_output.strip().splitlines()[-20:])
        _set_payment_result(key, result.returncode, tail)

        conn = get_connection()
        now = datetime.now().isoformat()
        if result.returncode == 0:
            # Parse CONFIRMATION_REF from output
            conf_ref = None
            for line in (result.stdout or "").splitlines():
                if line.startswith("CONFIRMATION_REF="):
                    conf_ref = line.split("=", 1)[1].strip()
            conn.execute(
                """UPDATE vendor_payments SET status = 'cleared', updated_at = ?,
                   payment_ref = COALESCE(?, payment_ref) WHERE id = ?""",
                (now, conf_ref, vendor_payment_id),
            )
            # Mark linked invoices as paid
            inv_rows = conn.execute(
                "SELECT invoice_number FROM vendor_payment_invoices WHERE payment_id = ?",
                (vendor_payment_id,),
            ).fetchall()
            for inv in inv_rows:
                _cross_link_invoice(conn, inv["invoice_number"],
                                    conf_ref or f"PORTAL-{vendor_payment_id}",
                                    datetime.now().strftime("%Y-%m-%d"))
            logger.info(f"Payment scraper {key} succeeded for vp#{vendor_payment_id}")
        else:
            conn.execute(
                "UPDATE vendor_payments SET status = 'failed', updated_at = ? WHERE id = ?",
                (now, vendor_payment_id),
            )
            logger.warning(f"Payment scraper {key} failed for vp#{vendor_payment_id}: exit {result.returncode}")

        conn.commit()
        conn.close()

    except subprocess.TimeoutExpired:
        _set_payment_result(key, -1, "Timed out after 5 minutes")
        conn = get_connection()
        conn.execute(
            "UPDATE vendor_payments SET status = 'failed', updated_at = ? WHERE id = ?",
            (datetime.now().isoformat(), vendor_payment_id),
        )
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Payment scraper {key} error: {e}")
        _set_payment_result(key, -1, str(e))
        try:
            conn = get_connection()
            conn.execute(
                "UPDATE vendor_payments SET status = 'failed', updated_at = ? WHERE id = ?",
                (datetime.now().isoformat(), vendor_payment_id),
            )
            conn.commit()
            conn.close()
        except Exception:
            pass
    finally:
        _clear_payment_running(key)


@payment_bp.route("/api/payments/pay-portal", methods=["POST"])
def api_pay_portal():
    """Create a processing payment and spawn scraper to pay via vendor portal."""
    data = request.get_json(silent=True) or {}
    vendor_name = data.get("vendor_name")
    invoice_ids = data.get("invoice_ids", [])
    amounts = data.get("amounts_per_invoice", [])

    if not vendor_name or not invoice_ids:
        return jsonify({"error": "vendor_name and invoice_ids required"}), 400

    # Validate vendor has portal enabled
    conn = get_connection()
    vbp = conn.execute(
        "SELECT portal_pay_enabled FROM vendor_bill_pay WHERE vendor_name = ?",
        (vendor_name,),
    ).fetchone()
    if not vbp or not vbp["portal_pay_enabled"]:
        conn.close()
        return jsonify({"error": "Portal pay not enabled for this vendor"}), 400

    # Resolve scraper key
    scraper_key = _VENDOR_KEY_MAP.get(vendor_name)
    if not scraper_key or scraper_key not in _PAYMENT_SCRAPER_REGISTRY:
        conn.close()
        return jsonify({"error": f"No payment scraper registered for {vendor_name}"}), 400

    scraper_info = _PAYMENT_SCRAPER_REGISTRY[scraper_key]

    # Check not already running
    state = _read_payment_scraper_state()
    if scraper_key in state.get("running", {}):
        conn.close()
        return jsonify({"error": f"Payment scraper for {vendor_name} is already running"}), 409

    # Calculate total
    total = sum(float(a) for a in amounts) if amounts else 0

    # Create vendor_payment with status=processing
    ref = f"PORTAL-{vendor_name}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    try:
        cur = conn.execute(
            """INSERT INTO vendor_payments
               (vendor, payment_date, payment_ref, payment_method, payment_total,
                status, source, created_at)
               VALUES (?, ?, ?, 'portal', ?, 'processing', 'portal', ?)""",
            (vendor_name, datetime.now().strftime("%Y-%m-%d"), ref, total,
             datetime.now().isoformat()),
        )
        vp_id = cur.lastrowid

        # Link invoices
        for i, inv_id in enumerate(invoice_ids):
            applied = float(amounts[i]) if i < len(amounts) else 0
            inv_row = conn.execute(
                "SELECT invoice_number, invoice_date, due_date FROM scanned_invoices WHERE id = ?",
                (inv_id,),
            ).fetchone()
            if inv_row:
                conn.execute(
                    """INSERT INTO vendor_payment_invoices
                       (payment_id, invoice_number, invoice_date, due_date, amount_paid)
                       VALUES (?, ?, ?, ?, ?)""",
                    (vp_id, inv_row["invoice_number"], inv_row["invoice_date"],
                     inv_row["due_date"], applied),
                )

        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({"error": f"Failed to create payment record: {e}"}), 500

    # Write payment_request.json for scraper
    request_dir = scraper_info["script_dir"]
    os.makedirs(request_dir, exist_ok=True)
    inv_details = []
    for i, inv_id in enumerate(invoice_ids):
        inv_row = conn.execute(
            "SELECT invoice_number, total, due_date FROM scanned_invoices WHERE id = ?",
            (inv_id,),
        ).fetchone()
        if inv_row:
            inv_details.append({
                "invoice_number": inv_row["invoice_number"],
                "amount": float(amounts[i]) if i < len(amounts) else float(inv_row["total"] or 0),
                "due_date": inv_row["due_date"],
            })
    conn.close()

    request_file = os.path.join(request_dir, "payment_request.json")
    with open(request_file, "w") as f:
        json.dump({
            "vendor_payment_id": vp_id,
            "vendor_name": vendor_name,
            "total": total,
            "invoices": inv_details,
            "requested_at": datetime.now().isoformat(),
        }, f, indent=2)

    # Spawn background thread
    t = threading.Thread(
        target=_run_payment_scraper_bg,
        args=(scraper_key, vp_id, scraper_info),
        daemon=True,
    )
    t.start()

    logger.info(f"Portal payment initiated: {vendor_name} ${total:.2f}, vp#{vp_id}")
    return jsonify({"status": "ok", "vendor_payment_id": vp_id, "payment_ref": ref})


@payment_bp.route("/api/payments/scraper-status", methods=["GET"])
def api_payment_scraper_status():
    """Return running/results from payment scraper state file."""
    state = _read_payment_scraper_state()
    return jsonify(state)


@payment_bp.route("/api/payments/scraper-log/<key>", methods=["GET"])
def api_payment_scraper_log(key):
    """Return full log from last payment scraper run."""
    log_path = os.path.join(_PAYMENT_LOG_DIR, f"payment_{key}.log")
    if not os.path.exists(log_path):
        return jsonify({"error": "No log found"}), 404
    with open(log_path) as f:
        return jsonify({"key": key, "log": f.read()})
