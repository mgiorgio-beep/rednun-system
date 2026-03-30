"""
Weekly Excel Export — Red Nun Analytics
Generates a comprehensive weekly Excel report combining:
  - Toast POS: Revenue, orders, discounts, server performance
  - 7shifts: Full payroll labor (hourly + salaried), by role
  - MarginEdge: COGS, vendor spending, purchasing
"""

import os
import logging
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from data_store import get_connection
from analytics import get_daily_revenue, get_server_performance

try:
    from sevenshifts_client import get_labor_for_report
    SEVENSHIFTS = True
except ImportError:
    SEVENSHIFTS = False

logger = logging.getLogger(__name__)

# Styles
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL_RED = PatternFill("solid", fgColor="C43B3B")
HEADER_FILL_BLUE = PatternFill("solid", fgColor="4A7FD4")
HEADER_FILL_DARK = PatternFill("solid", fgColor="2D3142")
SUBHEAD_FONT = Font(name="Arial", bold=True, size=10)
SUBHEAD_FILL = PatternFill("solid", fgColor="F0F0F0")
BODY_FONT = Font(name="Arial", size=10)
MONEY_FMT = '$#,##0'
MONEY2_FMT = '$#,##0.00'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD")
)
CENTER = Alignment(horizontal="center")
LOCATION_NAMES = {"dennis": "Dennis Port", "chatham": "Chatham"}
LOCATION_COLORS = {"dennis": HEADER_FILL_RED, "chatham": HEADER_FILL_BLUE}
DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _style_header(ws, row, max_col, fill=HEADER_FILL_DARK):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER


def _style_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).font = BODY_FONT
        ws.cell(row=row, column=col).border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 25)


def _add_revenue_sheet(wb, start_date, end_date, location=None):
    ws = wb.active
    ws.title = "Revenue Summary"

    # Per-location daily revenue
    all_days = []
    locations = [location] if location else ["dennis", "chatham"]
    for loc in locations:
        daily = get_daily_revenue(loc, start_date, end_date)
        for d in daily:
            d["location"] = loc
            try:
                dt = datetime.strptime(d["business_date"], "%Y%m%d")
                d["day_name"] = DAY_NAMES[dt.weekday()]
            except (ValueError, TypeError):
                d["day_name"] = ""
        all_days.extend(daily)

    all_days.sort(key=lambda x: (x["business_date"], x["location"]))

    headers = ["Date", "Day", "Location", "Net Revenue", "Tax", "Tips",
               "Discounts", "Orders", "Avg Check"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for d in all_days:
        avg_check = d["net_revenue"] / d["order_count"] if d["order_count"] else 0
        row = [
            d["business_date"], d["day_name"],
            "Dennis Port" if d["location"] == "dennis" else "Chatham",
            d["net_revenue"], d.get("tax_amount", d.get("tax", 0)),
            d.get("tip_amount", d.get("tips", 0)),
            d.get("discount_amount", d.get("discounts", 0)),
            d["order_count"], round(avg_check, 2),
        ]
        ws.append(row)
        r = ws.max_row
        _style_row(ws, r, len(headers))
        ws.cell(r, 4).number_format = MONEY_FMT
        ws.cell(r, 5).number_format = MONEY_FMT
        ws.cell(r, 6).number_format = MONEY_FMT
        ws.cell(r, 7).number_format = MONEY_FMT
        ws.cell(r, 9).number_format = MONEY2_FMT

    # Totals row
    last = ws.max_row
    ws.append([
        "", "", "TOTAL",
        f"=SUM(D2:D{last})", f"=SUM(E2:E{last})", f"=SUM(F2:F{last})",
        f"=SUM(G2:G{last})", f"=SUM(H2:H{last})",
        f"=D{last+1}/H{last+1}",
    ])
    r = ws.max_row
    for col in range(1, len(headers) + 1):
        ws.cell(r, col).font = Font(name="Arial", bold=True, size=10)
    ws.cell(r, 4).number_format = MONEY_FMT
    ws.cell(r, 9).number_format = MONEY2_FMT

    _auto_width(ws)
    return ws


def _add_labor_sheet(wb, start_date, end_date, location=None):
    ws = wb.create_sheet("Labor — 7shifts")

    if not SEVENSHIFTS:
        ws.append(["7shifts not available — install sevenshifts_client.py"])
        return ws

    headers = ["Location", "Role", "Hours", "Pay", "Tips", "Shifts"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    grand_pay = 0
    grand_hours = 0

    locations = [(location, LOCATION_NAMES.get(location, location))] if location else [("dennis", "Dennis Port"), ("chatham", "Chatham")]

    for loc, loc_label in locations:
        labor = get_labor_for_report(loc, start_date, end_date)
        roles = labor.get("by_role", [])

        for r in roles:
            if r["role"] == "Management (Salary)":
                continue
            ws.append([loc_label, r["role"], r["hours"], r["pay"], r.get("tips", 0), r["shifts"]])
            row = ws.max_row
            _style_row(ws, row, len(headers))
            ws.cell(row, 3).number_format = '#,##0.0'
            ws.cell(row, 4).number_format = MONEY_FMT
            ws.cell(row, 5).number_format = MONEY_FMT

        # Salaried line
        sal = labor.get("salaried_cost", 0)
        if sal > 0:
            ws.append([loc_label, "Management (Salary)", "", sal, "", ""])
            row = ws.max_row
            _style_row(ws, row, len(headers))
            ws.cell(row, 4).number_format = MONEY_FMT

        grand_pay += labor.get("total_labor_cost", 0)
        grand_hours += labor.get("total_hours", 0)

        # Location subtotal
        ws.append(["", f"  {loc_label} Total", labor.get("total_hours", 0),
                    labor.get("total_labor_cost", 0), labor.get("total_tips", 0), ""])
        row = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row, col).font = Font(name="Arial", bold=True, size=10)
            ws.cell(row, col).fill = SUBHEAD_FILL
        ws.cell(row, 3).number_format = '#,##0.0'
        ws.cell(row, 4).number_format = MONEY_FMT
        ws.cell(row, 5).number_format = MONEY_FMT
        ws.append([])  # blank row

    # Employee detail section
    ws.append([])
    ws.append(["Employee Detail"])
    row = ws.max_row
    ws.cell(row, 1).font = Font(name="Arial", bold=True, size=11)

    emp_headers = ["Location", "Employee", "Roles", "Hours", "Pay", "Tips", "Salaried"]
    ws.append(emp_headers)
    _style_header(ws, ws.max_row, len(emp_headers))

    for loc, loc_label in locations:
        labor = get_labor_for_report(loc, start_date, end_date)
        for e in labor.get("by_employee", []):
            ws.append([
                loc_label, e["name"], e["roles"],
                e["hours"], e["pay"], e["tips"],
                "Yes" if e["salaried"] else "",
            ])
            row = ws.max_row
            _style_row(ws, row, len(emp_headers))
            ws.cell(row, 4).number_format = '#,##0.0'
            ws.cell(row, 5).number_format = MONEY_FMT
            ws.cell(row, 6).number_format = MONEY_FMT

    _auto_width(ws)
    return ws


def _add_cogs_sheet(wb, start_date, end_date, location=None):
    ws = wb.create_sheet("COGS — MarginEdge")
    conn = get_connection()

    # Convert YYYYMMDD to YYYY-MM-DD
    sd = f"{start_date[:4]}-{start_date[4:6]}-{start_date[6:]}"
    ed = f"{end_date[:4]}-{end_date[4:6]}-{end_date[6:]}"

    loc_clause = "AND location = ?" if location else ""
    base_params = (sd, ed, location) if location else (sd, ed)

    # COGS by category per location
    headers = ["Location", "Category", "Total Cost", "Invoices"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    # Vendor-to-category mapping (matches marginedge_sync.py)
    VENDOR_HINTS = {
        "LIQUOR": ["southern glazer", "l. knife", "martignetti", "atlantic beverage", "horizon beverage"],
        "BEER": ["colonial wholesale", "craft collective", "cape cod beer"],
        "FOOD": ["us foods", "reinhart", "performance food", "chefs warehouse", "cape fish", "sysco"],
        "NON_COGS": ["cintas", "unifirst", "cozzini", "rooter", "dennisport village",
                      "caron group", "robert b. our", "marginedge"],
    }

    def _vendor_category(name):
        lower = (name or "").lower()
        for cat, patterns in VENDOR_HINTS.items():
            for p in patterns:
                if p in lower:
                    return cat
        return "OTHER"

    try:
        invoices = conn.execute(f"""
            SELECT location, vendor_name, SUM(order_total) as total_cost, COUNT(*) as cnt
            FROM me_invoices
            WHERE invoice_date >= ? AND invoice_date <= ? {loc_clause}
            GROUP BY location, vendor_name
            ORDER BY location, total_cost DESC
        """, base_params).fetchall()

        # Aggregate by category
        cat_totals = {}
        for inv in invoices:
            loc = inv["location"]
            cat = _vendor_category(inv["vendor_name"])
            key = (loc, cat)
            if key not in cat_totals:
                cat_totals[key] = {"location": loc, "category_type": cat, "total_cost": 0, "invoice_count": 0}
            cat_totals[key]["total_cost"] += inv["total_cost"]
            cat_totals[key]["invoice_count"] += inv["cnt"]

        for r in sorted(cat_totals.values(), key=lambda x: (x["location"], -x["total_cost"])):
            loc_label = "Dennis Port" if r["location"] == "dennis" else "Chatham"
            ws.append([loc_label, r["category_type"], r["total_cost"], r["invoice_count"]])
            row = ws.max_row
            _style_row(ws, row, len(headers))
            ws.cell(row, 3).number_format = MONEY_FMT
    except Exception as e:
        ws.append([f"COGS data not available: {e}"])

    # Vendor spending section
    ws.append([])
    ws.append(["Vendor Spending (Weekly)"])
    row = ws.max_row
    ws.cell(row, 1).font = Font(name="Arial", bold=True, size=11)

    vendor_headers = ["Location", "Vendor", "Category", "Total Spend", "Invoices"]
    ws.append(vendor_headers)
    _style_header(ws, ws.max_row, len(vendor_headers))

    try:
        vendors = conn.execute(f"""
            SELECT location, vendor_name,
                   SUM(order_total) as total_cost,
                   COUNT(*) as invoice_count
            FROM me_invoices
            WHERE invoice_date >= ? AND invoice_date <= ? {loc_clause}
            GROUP BY location, vendor_name
            ORDER BY location, total_cost DESC
        """, base_params).fetchall()

        for v in vendors:
            loc_label = "Dennis Port" if v["location"] == "dennis" else "Chatham"
            ws.append([loc_label, v["vendor_name"], "",
                       v["total_cost"], v["invoice_count"]])
            row = ws.max_row
            _style_row(ws, row, len(vendor_headers))
            ws.cell(row, 4).number_format = MONEY_FMT
    except Exception as e:
        ws.append([f"Vendor data not available: {e}"])

    conn.close()
    _auto_width(ws)
    return ws


def _add_server_sheet(wb, start_date, end_date, location=None):
    ws = wb.create_sheet("Server Performance")

    headers = ["Rank", "Server", "Location", "Total Sales", "Orders",
               "Avg Check", "Tips", "Avg Tip"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    servers = get_server_performance(location, start_date, end_date)
    for i, s in enumerate(servers or [], 1):
        loc_label = "Dennis Port" if s.get("location") == "dennis" else "Chatham"
        avg_check = s["total_sales"] / s["order_count"] if s["order_count"] else 0
        tips = s.get("tips", s.get("tip_amount", 0)) or 0
        avg_tip = tips / s["order_count"] if s["order_count"] else 0
        ws.append([
            i, s["server_name"], loc_label,
            s["total_sales"], s["order_count"],
            round(avg_check, 2), tips, round(avg_tip, 2),
        ])
        row = ws.max_row
        _style_row(ws, row, len(headers))
        ws.cell(row, 4).number_format = MONEY_FMT
        ws.cell(row, 6).number_format = MONEY2_FMT
        ws.cell(row, 7).number_format = MONEY_FMT
        ws.cell(row, 8).number_format = MONEY2_FMT

    _auto_width(ws)
    return ws


def generate_weekly_excel(start_date, end_date, output_dir=None, location=None):
    """Generate comprehensive weekly Excel report, optionally filtered by location."""
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(__file__), "exports")
    os.makedirs(output_dir, exist_ok=True)

    loc_suffix = f"_{location}" if location else ""
    filename = f"RedNun_Weekly_{start_date}_to_{end_date}{loc_suffix}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()

    _add_revenue_sheet(wb, start_date, end_date, location)
    _add_labor_sheet(wb, start_date, end_date, location)
    _add_cogs_sheet(wb, start_date, end_date, location)
    _add_server_sheet(wb, start_date, end_date, location)

    wb.save(filepath)
    logger.info(f"Excel report saved: {filepath}")
    return filepath


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    if len(sys.argv) == 3:
        start, end = sys.argv[1], sys.argv[2]
    else:
        today = datetime.now().date()
        last_monday = today - timedelta(days=today.weekday() + 7)
        last_sunday = last_monday + timedelta(days=6)
        start = last_monday.strftime("%Y%m%d")
        end = last_sunday.strftime("%Y%m%d")

    path = generate_weekly_excel(start, end)
    print(f"Report generated: {path}")
